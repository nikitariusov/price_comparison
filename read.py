import openpyxl


def read_file(file):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column

    clean_file(sheet)

    name_cols = ['Название', 'Бренд', 'Ссылка КТУ', 'Цена КТУ',
                 'Конкурент 1', 'Ссылка конкурента 1', 'Цена конкурента 1',
                 'Конкурент 2', 'Ссылка конкурента 2', 'Цена конкурента 2',
                 'Конкурент 3', 'Ссылка конкурента 3', 'Цена конкурента 3',
                 'Конкурент 4', 'Ссылка конкурента 4', 'Цена конкурента 4',
                 'Конкурент 5', 'Ссылка конкурента 5', 'Цена конкурента 5']

    contents = []
    for row in sheet.iter_rows(min_row=2, max_col=cols, max_row=rows):
        dic = {}
        items = []
        for cell in row:
            items.append(cell.value)

        for i in range(0, len(items)):
            dic[name_cols[i]] = items[i]

        contents.append(dic)
    return contents


def clean_file(sheet):
    list_cols_name = ['D', 'G', 'J', 'M', 'P', 'S']
    for i in list_cols_name:
        for j in range(2, sheet.max_row):
            sheet[f'{i}{j}'] = ''


if __name__ == '__main__':
    file = 'test/prods2.xlsx'
    print(read_file(file))
