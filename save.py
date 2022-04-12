import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime, date
from colorama import init, Fore


def try_save(sheet, i, j, cols_name, price):
    try:
        if i[price]:
            sheet[f'{cols_name}{j}'] = i[price]
            sheet[f'{cols_name}{j}'].font = font
            try:
                if i['Цена КТУ'] < i[price]:
                    sheet[f'{cols_name}{j}'].fill = upper_fill
                elif i['Цена КТУ'] > i[price]:
                    sheet[f'{cols_name}{j}'].fill = lower_fill
            except:
                pass
    except KeyError:
        pass


font = Font(name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')

upper_fill = PatternFill(fill_type='solid',
                         start_color='00FF0000',
                         end_color='00FF0000')

lower_fill = PatternFill(fill_type='solid',
                         start_color='0000FF00',
                         end_color='0000FF00')


def save_file(data, file):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb.active

    j = 2
    for i in data:
        try_save(sheet, i, j, 'D', 'Цена КТУ')
        try_save(sheet, i, j, 'G', 'Цена конкурента 1')
        try_save(sheet, i, j, 'J', 'Цена конкурента 2')
        try_save(sheet, i, j, 'M', 'Цена конкурента 3')
        try_save(sheet, i, j, 'P', 'Цена конкурента 4')
        try_save(sheet, i, j, 'S', 'Цена конкурента 5')
        j += 1

    filename = f'Проверен-{date.today()}.xlsx'
    wb.save(filename)

    init(autoreset=True)  # запускаем колораму
    print(Fore.GREEN + f'''[INFO]\tCохранен файл {filename}''')
    print('''\nВыход - Enter''')


if __name__ == '__main__':
    print(date.today())
    init(autoreset=True)
    print(Fore.GREEN + f'''[INFO]\tCохранен файл ''')
    print('''\nВыход - Enter''')
